# -*- coding: utf-8 -*-
from odoo import http
from odoo.http import request
import json
import ast
import pandas as pd
import io
import base64
import logging
import markdown
from datetime import datetime
import re
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from langchain_deepseek import ChatDeepSeek
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser

_logger = logging.getLogger(__name__)

# -------------------------------------------------------------
#   FIXED PATH TO YOUR MIS MODULES
# -------------------------------------------------------------
MIS_MODULE_PATH = "/home/moses/workspace/odoo_mis"


# -------------------------------------------------------------
#   CLEAN DATETIME HELPER
# -------------------------------------------------------------
def clean_datetime_repr(data_list):
    if not data_list or not isinstance(data_list, list):
        return data_list
    cleaned = []
    for rec in data_list:
        if isinstance(rec, dict):
            new = {}
            for k, v in rec.items():
                if isinstance(v, str) and "datetime.datetime(" in v:
                    try:
                        v = eval(v)
                    except Exception:
                        pass
                new[k] = v
            cleaned.append(new)
        else:
            cleaned.append(rec)
    return cleaned

def get_mis_module_names():
    """
    Returns a set of module names whose Python models come from /odoo_mis/
    """
    registry = request.env.registry
    mis_path = "/home/moses/workspace/odoo_mis/"

    mis_modules = set()

    for model_name, model_obj in registry.models.items():
        try:
            file_path = getattr(model_obj, "__file__", None)
            if file_path and file_path.startswith(mis_path):
                mis_modules.add(model_obj._module)
        except Exception:
            continue

    return sorted(mis_modules)



# -------------------------------------------------------------
#   BUILD MIS SCHEMA (AI SQL SCHEMA)
# -------------------------------------------------------------
def get_mis_schema():
    """
    Build schema only for required MIS models and replace dots with underscores in the result.
    Example:
        education.student → education_student
    """
    env = request.env
    registry = env["ir.model"].sudo()

    allowed_prefixes = (
        "education.",
        "student.",
        "school.",
        "exam.",
        "results.",
    )

    all_models = registry.search([])
    filtered_models = all_models.filtered(
        lambda m: any(m.model.startswith(prefix) for prefix in allowed_prefixes)
    )

    schema = {}

    for model in filtered_models:
        model_name = model.model                       # e.g. education.student
        clean_model_name = model_name.replace(".", "_")  # → education_student

        fields = env[model_name].sudo().fields_get()

        field_schema = {}
        for fname, finfo in fields.items():
            field_type = finfo.get("type")
            entry = {"type": field_type}

            if field_type in ("many2one", "one2many", "many2many"):
                entry["relation"] = finfo.get("relation")

            field_schema[fname] = entry

        # store using cleaned model name
        schema[clean_model_name] = {
            "fields": field_schema
        }

    return schema




# -------------------------------------------------------------
#   AI — ICSE SCHOOL SYSTEM MESSAGE
# -------------------------------------------------------------
def icse_system_prompt():
    """
    Returns a system prompt optimized for ICSE schools in India.
    Personalized for:
    Mary Immaculate English Medium School (MIS Basirhat)
    """
    return (
        "You are a highly specialized data analysis assistant for "
        "Mary Immaculate English Medium School, Basirhat (ICSE), India.\n\n"

        "You understand:\n"
        "- ICSE syllabus, classes LKG–X\n"
        "- ICSE evaluation patterns: Theory, Practical, Project, Internal Assessment\n"
        "- ICSE grading system: A1, A2, B1, B2, C1, C2, D, E\n"
        "- Attendance rules, student records, divisions, roll numbers\n"
        "- Homework, events, notices, and academic year workflows\n"
        "- School fees, dues, concessions, monthly and term-wise fees\n"
        "- Timetables, teachers, subjects, exams, marks entry\n"
        "- Indian culture, ₹ currency, and DD-MM-YYYY date format\n\n"

        "When explaining results:\n"
        "- Be clear, simple, and school-friendly\n"
        "- Avoid technical jargon\n"
        "- Use descriptive insights helpful for teachers and admins\n"
        "- Format money in ₹ (e.g., ₹12,350)\n"
        "- Format dates as DD-MM-YYYY\n"
    )


# -------------------------------------------------------------
#   EXTRACT SQL FROM AI RESPONSE
# -------------------------------------------------------------
def extract_sql(text):
    # ```sql ... ```
    m = re.search(r"```sql\s*([\s\S]*?)```", text)
    if m:
        return m.group(1).strip()

    # ``` ... ```
    m = re.search(r"```([\s\S]*?)```", text)
    if m:
        return m.group(1).strip()

    # SELECT ... ;
    m = re.search(r"(SELECT[\s\S]+?;)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip()

    return text.strip()

# =====================================================================
#   PART 2 — MAIN AI SEARCH CONTROLLER (MIS ONLY + ICSE OPTIMIZED)
# =====================================================================

class AiSearchController(http.Controller):

    @http.route('/ai/search', type='http', auth="user", website=True, csrf=True)
    def ai_search(self, **post):

        query = post.get("chatInput", "").strip()
        if not query:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "result_data": None,
                "user_query": "",
                "search_history": [],
                "error_message": "Please enter a question."
            })

        _logger.info("🔥 User Query: %s", query)

        # -------------------------------------------------------------
        #   LOAD MIS SCHEMA (ONLY MODULES FROM /home/moses/workspace/odoo_mis)
        # -------------------------------------------------------------
        schema = get_mis_schema()
        if not schema:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "result_data": None,
                "user_query": query,
                "search_history": [],
                "error_message": "No MIS schema found. Check MIS module installation."
            })

        _logger.info("📘 Loaded MIS Schema (tables=%s)", len(schema))

        # -------------------------------------------------------------
        #   AI CLIENT (DeepSeek)
        # -------------------------------------------------------------
        api_key = request.env['ir.config_parameter'].sudo().get_param('mis_ai_analysis.openai_api_key')
        print('ffffffffffffff',api_key)
        if not api_key:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "result_data": None,
                "user_query": query,
                "search_history": [],
                "error_message": "DeepSeek API key not configured in System Parameters."
            })

        llm = ChatDeepSeek(model="deepseek-chat", temperature=0, api_key=api_key)

        # -------------------------------------------------------------
        #   1) REPHRASE USER QUERY (ICSE SCHOOL CONTEXT)
        # -------------------------------------------------------------
        try:
            prompt_rephrase = ChatPromptTemplate.from_messages([
                ("system", icse_system_prompt()),
                ("user",
                 "Rewrite this question clearly for a school-management context.\n"
                 "Do NOT write SQL.\n"
                 "Question: {q}")
            ])
            chain = prompt_rephrase | llm | StrOutputParser()
            rephrased_query = chain.invoke({"q": query}).strip()

        except Exception as e:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "result_data": None,
                "user_query": query,
                "search_history": [],
                "error_message": "AI failed during query rephrasing: %s" % e
            })

        _logger.info("✏ Rephrased: %s", rephrased_query)

        # -------------------------------------------------------------
        #   2) CHOOSE RELEVANT TABLES FROM MIS SCHEMA
        # -------------------------------------------------------------


        print('SWWWWWWWWWWWWWWWWWWWWWW',schema)
        try:
            prompt_schema = ChatPromptTemplate.from_messages([
                ("system",
                 icse_system_prompt() +
                 "\nSelect ONLY the relevant MIS tables and fields for answering the question."),
                ("user",
                 "User question: {rq}\n\nHere is the MIS schema:\n{schema_json}")
            ])
            relevant_schema_resp = (prompt_schema | llm).invoke({
                "rq": rephrased_query,
                "schema_json": json.dumps(schema, indent=2)
            })
            relevant_schema = relevant_schema_resp.content.strip()
        except Exception as e:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "result_data": None,
                "user_query": query,
                "search_history": [],
                "error_message": "AI failed during schema selection: %s" % e
            })

        _logger.info("🗂 Relevant schema chosen")

        # -------------------------------------------------------------
        #   3) GENERATE SQL FROM MIS SCHEMA
        # -------------------------------------------------------------
        try:
            prompt_sql = ChatPromptTemplate.from_messages([
                ("system",
                 icse_system_prompt() +
                 "\nGenerate ONLY a PostgreSQL SELECT query.\n"
                 "Rules:\n"
                 "- Output ONLY SQL (no explanations)\n"
                 "- Use ILIKE with % for text matching\n"
                 "- Use date format YYYY-MM-DD in SQL\n"
                 "- Select only columns from relevant MIS schema\n"
                 "- Never invent columns or tables"),
                ("user",
                 "Rephrased question: {rq}\n\nRelevant schema:\n{rel_schema}\n\nWrite SQL:")
            ])
            sql_resp = (prompt_sql | llm).invoke({
                "rq": rephrased_query,
                "rel_schema": relevant_schema
            })
            sql_raw = sql_resp.content.strip()
            sql_clean = extract_sql(sql_raw)

        except Exception as e:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "result_data": None,
                "user_query": query,
                "search_history": [],
                "error_message": "AI failed during SQL generation: %s" % e
            })

        _logger.info("📝 AI SQL:\n%s", sql_clean)

        # -------------------------------------------------------------
        #   4) EXECUTE SQL ON THE SAME ODOO DB
        # -------------------------------------------------------------
        cr = request.env.cr
        rows = []
        last_err = None
        attempts = 4
        sql_to_try = sql_clean

        for i in range(attempts):
            try:
                _logger.info("▶ SQL Attempt %s: %s", i + 1, sql_to_try)
                cr.execute(sql_to_try)
                rows = cr.dictfetchall()
                break  # SUCCESS
            except Exception as e:
                request.env.cr.rollback()
                last_err = str(e)
                _logger.warning("❌ SQL Error (attempt %s): %s", i + 1, last_err)

                if i == attempts - 1:
                    return request.render("mis_ai_analysis.ai_ama_chat_window", {
                        "result_data": None,
                        "user_query": query,
                        "search_history": [],
                        "error_message": "SQL failed even after corrections:\n%s" % last_err
                    })

                # Ask AI to fix SQL
                try:
                    fix_prompt = ChatPromptTemplate.from_messages([
                        ("system",
                         icse_system_prompt() +
                         "\nFix the SQL query based on the PostgreSQL error."),
                        ("user",
                         "SQL:\n{sql}\n\nError:\n{err}\n\nMIS Schema:\n{schema_json}\n\nCorrect the SQL:")
                    ])
                    fix_resp = (fix_prompt | llm).invoke({
                        "sql": sql_to_try,
                        "err": last_err,
                        "schema_json": json.dumps(schema, indent=2)
                    })
                    sql_to_try = extract_sql(fix_resp.content.strip())
                except Exception as e_fix:
                    return request.render("mis_ai_analysis.ai_ama_chat_window", {
                        "result_data": None,
                        "user_query": query,
                        "search_history": [],
                        "error_message": "AI failed while fixing SQL: %s" % e_fix
                    })

        # -------------------------------------------------------------
        #   5) AI EXPLANATION
        # -------------------------------------------------------------
        try:
            explain_prompt = ChatPromptTemplate.from_messages([
                ("system",
                 icse_system_prompt() +
                 "\nExplain the SQL result in clear, school-friendly English."),
                ("user",
                 "User question: {uq}\n\nResults:\n{rows}")
            ])
            explanation = (explain_prompt | llm).invoke({
                "uq": query,
                "rows": rows
            }).content.strip()

            explanation_html = markdown.markdown(explanation)

        except Exception as e:
            explanation_html = "<p>AI explanation unavailable: %s</p>" % e

        # -------------------------------------------------------------
        #   6) BUILD CHART DATA FOR FRONTEND (if numeric + text columns exist)
        # -------------------------------------------------------------
        df = pd.DataFrame(rows)
        chart_data = None

        if not df.empty:
            try:
                text_cols = [c for c in df.columns if df[c].dtype == "object"]
                num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

                if text_cols and num_cols:
                    x_col = text_cols[0]
                    df["x_axis"] = df[x_col].astype(str)

                    chart_data = {
                        "x": df["x_axis"].tolist(),
                        "series": []
                    }
                    for nc in num_cols:
                        chart_data["series"].append({
                            "name": nc,
                            "data": df[nc].fillna(0).tolist()
                        })

            except Exception as e:
                _logger.error("Chart build error: %s", e)

        # -------------------------------------------------------------
        #   7) STORE SEARCH HISTORY
        # -------------------------------------------------------------
        hist = request.env['ai.user.queries'].sudo().create({
            'query_text': query,
            'result_from_db': rows,
            'generated_query': sql_to_try,
            'create_date': datetime.now(),
        })

        user = request.env.user
        search_history = request.env['ai.user.queries'].sudo().search([], limit=50)

        # -------------------------------------------------------------
        #   8) RENDER UI TEMPLATE
        # -------------------------------------------------------------
        return request.render("mis_ai_analysis.ai_ama_chat_window", {
            "result_data": rows,
            "user_query": query,
            "ai_explanation_html": explanation_html,
            "chart_data": chart_data,
            "search_history": search_history,
            "download_xlsx_url": f"/ai/download_xlsx/{hist.id}",
            "active_tab": "preview",
        })


# =====================================================================
#   PART 3 — XLSX DOWNLOAD (Data + AI Explanation + Chart)
# =====================================================================

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage

class AiXlsxController(http.Controller):

    @http.route('/ai/download_xlsx/<int:search_log_id>', type='http', auth="user")
    def ai_download_xlsx(self, search_log_id, **kw):

        log = request.env["ai.search.log"].sudo().browse(search_log_id)
        if not log:
            return request.not_found()

        rows = log.result_from_db or []
        explanation = log.ai_explanation or "No explanation available."

        # Build Excel Workbook
        wb = Workbook()

        # =====================================================
        #  SHEET 1 : DATA
        # =====================================================
        ws_data = wb.active
        ws_data.title = "Data"

        if rows:
            # Write header
            headers = list(rows[0].keys())
            for col_i, h in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col_i, value=h.replace("_", " ").title())
                cell.font = Font(bold=True)

            # Write rows
            for r_i, rec in enumerate(rows, start=2):
                for c_i, h in enumerate(headers, 1):
                    val = rec.get(h)
                    ws_data.cell(row=r_i, column=c_i, value=val)

        # Auto column width
        for column_cells in ws_data.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws_data.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 60)

        # =====================================================
        #  SHEET 2 : AI EXPLANATION
        # =====================================================
        ws_exp = wb.create_sheet("AI Explanation")
        ws_exp["A1"] = "User Query:"
        ws_exp["A1"].font = Font(bold=True)

        ws_exp["B1"] = log.query_text
        ws_exp["A3"] = "AI Explanation:"
        ws_exp["A3"].font = Font(bold=True)

        ws_exp["B3"] = explanation
        ws_exp["B3"].alignment = Alignment(wrap_text=True)

        ws_exp.column_dimensions["A"].width = 25
        ws_exp.column_dimensions["B"].width = 100

        # =====================================================
        #  SHEET 3 : CHART (Optional)
        # =====================================================
        try:
            chart_data = rows

            if rows and isinstance(rows, list):
                df = pd.DataFrame(rows)
                num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
                text_cols = [c for c in df.columns if df[c].dtype == "object"]

                if num_cols and text_cols:
                    x_col = text_cols[0]
                    df["x_axis"] = df[x_col].astype(str)

                    # Create plot
                    fig, ax = plt.subplots(figsize=(10, 4))

                    for y in num_cols:
                        ax.plot(df["x_axis"], df[y], marker="o", label=y)

                    plt.xticks(rotation=45, ha="right")
                    plt.tight_layout()

                    buf = io.BytesIO()
                    plt.savefig(buf, format="png", dpi=150)
                    buf.seek(0)

                    ws_chart = wb.create_sheet("Chart")
                    img = XLImage(buf)
                    img.anchor = "A1"
                    ws_chart.add_image(img)
                    plt.close()

        except Exception as e:
            _logger.warning("Chart creation failed for XLSX: %s", e)

        # =====================================================
        #  RETURN EXCEL
        # =====================================================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return request.make_response(
            output.getvalue(),
            headers=[
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="AI_School_Report_{search_log_id}.xlsx"')
            ]
        )

