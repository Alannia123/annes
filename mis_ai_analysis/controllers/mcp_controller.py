# -*- coding: utf-8 -*-
from odoo import http
from odoo.http import request
import json
import requests
import logging
import markdown
from datetime import datetime
import pandas as pd

_logger = logging.getLogger(__name__)
# -------------------------------------------------------------
# LOAD MCP CONFIG FROM ir.config_parameter
# -------------------------------------------------------------
def get_mcp_config():
    ICP = request.env['ir.config_parameter'].sudo()

    mcp_url = ICP.get_param('mis_ai_analysis.mcp_server_url')
    mcp_secret = ICP.get_param('mis_ai_analysis.mcp_internal_secret')
    print('ffgffg',mcp_url,mcp_secret)

    if not mcp_url:
        raise Exception("MCP Server URL not configured in System Parameters.")

    if not mcp_secret:
        raise Exception("MCP Internal Secret not configured in System Parameters.")

    return mcp_url.rstrip("/"), mcp_secret

# -------------------------------------------------------------
# BUILD MIS SCHEMA (same as before)
# -------------------------------------------------------------
def get_mis_schema():

    env = request.env
    registry = env["ir.model"].sudo()

    allowed_prefixes = (
        "education.",
        "student.",
        "school.",
        "exam.",
        "results.",
    )

    all_models = registry.search([])
    filtered_models = all_models.filtered(
        lambda m: any(m.model.startswith(prefix) for prefix in allowed_prefixes)
    )

    schema = {}

    for model in filtered_models:
        model_name = model.model
        clean_model_name = model_name.replace(".", "_")

        fields = env[model_name].sudo().fields_get()

        field_schema = {}
        for fname, finfo in fields.items():
            entry = {"type": finfo.get("type")}
            if finfo.get("relation"):
                entry["relation"] = finfo.get("relation")
            field_schema[fname] = entry

        schema[clean_model_name] = {"fields": field_schema}

    return schema


# =============================================================
# MCP BASED AI SEARCH CONTROLLER
# =============================================================

class AiSearchController(http.Controller):

    @http.route('/ai/search', type='http', auth="user", website=True, csrf=True)
    def ai_search(self, **post):

        query = post.get("chatInput", "").strip()
        if not query:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "error_message": "Please enter a question."
            })

        _logger.info("🔥 User Query: %s", query)

        # -------------------------------------------------------------
        # 1️⃣ LOAD MIS SCHEMA
        # -------------------------------------------------------------
        schema = get_mis_schema()

        if not schema:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "error_message": "MIS schema not found."
            })

        # -------------------------------------------------------------
        # 2️⃣ CALL MCP → GENERATE SQL
        # -------------------------------------------------------------
        try:
            mcp_url, mcp_secret = get_mcp_config()

            resp = requests.post(
                f"{mcp_url}/mcp/tools/odoo.ai.generate_sql",
                headers={"x-internal-secret": mcp_secret},
                json={
                    "user_query": query,
                    "schema": schema
                },
                timeout=60
            )

            if resp.status_code != 200:
                raise Exception(resp.text)

            data = resp.json()
            sql_query = data["output"]["sql"]
            rephrased_query = data["output"]["rephrased_query"]

        except Exception as e:
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "error_message": f"MCP SQL generation failed: {e}"
            })

        _logger.info("📝 SQL from MCP:\n%s", sql_query)

        # -------------------------------------------------------------
        # 3️⃣ EXECUTE SQL LOCALLY (SECURE)
        # -------------------------------------------------------------
        try:
            cr = request.env.cr
            cr.execute(sql_query)
            rows = cr.dictfetchall()
        except Exception as e:
            request.env.cr.rollback()
            return request.render("mis_ai_analysis.ai_ama_chat_window", {
                "error_message": f"SQL Execution failed: {e}"
            })

        # -------------------------------------------------------------
        # 4️⃣ CALL MCP → EXPLAIN RESULTS
        # -------------------------------------------------------------
        try:

            mcp_url, mcp_secret = get_mcp_config()

            explain_resp = requests.post(
                f"{mcp_url}/mcp/tools/odoo.ai.explain_results",
                headers={"x-internal-secret": mcp_secret.strip()},
                json={
                    "rows": rows,
                    "user_query": query,
                    "ai_query": sql_query
                },
                timeout=60
            )

            if explain_resp.status_code != 200:
                raise Exception(explain_resp.text)

            explanation = explain_resp.json()["output"]["explanation"]

        except Exception as e:
            explanation = f"AI explanation failed: {e}"

        explanation_html = markdown.markdown(explanation)

        # -------------------------------------------------------------
        # 5️⃣ BUILD CHART DATA (LOCAL)
        # -------------------------------------------------------------
        df = pd.DataFrame(rows)
        chart_data = None

        if not df.empty:
            text_cols = [c for c in df.columns if df[c].dtype == "object"]
            num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

            if text_cols and num_cols:
                x_col = text_cols[0]
                chart_data = {
                    "x": df[x_col].astype(str).tolist(),
                    "series": [
                        {
                            "name": nc,
                            "data": df[nc].fillna(0).tolist()
                        } for nc in num_cols
                    ]
                }

        # -------------------------------------------------------------
        # 6️⃣ SAVE SEARCH LOG
        # -------------------------------------------------------------
        exist_query = request.env['ai.user.queries'].sudo().search([('query_text', '=', query)])
        if not exist_query:
            hist = request.env['ai.user.queries'].sudo().create({
                'query_text': query,
                'result_from_db': rows,
                'generated_query': sql_query,
                'create_date': datetime.now(),
            })

        search_log = request.env['ai.search.log'].sudo().create({
                'query_text': query,
                'rephrased_text': rephrased_query,
                'result_text': rows,
                'generated_query': sql_query,
                'create_date': datetime.now(),
            })

        search_history = request.env['ai.user.queries'].sudo().search([], limit=50)

        # -------------------------------------------------------------
        # 7️⃣ RENDER UI
        # -------------------------------------------------------------
        return request.render("mis_ai_analysis.ai_ama_chat_window", {
            "result_data": rows,
            "user_query": query,
            "ai_explanation_html": explanation_html,
            "chart_data": chart_data,
            'search_log_id': search_log.id,
            "search_history": search_history,
            "download_xlsx_url": f"/ai/download_xlsx/{search_log.id}",
            "active_tab": "preview",
        })



# =====================================================================
#   PART 3 — XLSX DOWNLOAD (Data + AI Explanation + Chart)
# =====================================================================

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import io


class AiXlsxController(http.Controller):

    @http.route('/ai/download_xlsx/<int:search_log_id>', type='http', auth="user")
    def ai_download_xlsx(self, search_log_id, **kw):

        log = request.env["ai.search.log"].sudo().browse(search_log_id)
        if not log:
            return request.not_found()

        rows = log.result_from_db or []
        explanation = log.ai_explanation or "No explanation available."

        # Build Excel Workbook
        wb = Workbook()

        # =====================================================
        #  SHEET 1 : DATA
        # =====================================================
        ws_data = wb.active
        ws_data.title = "Data"

        if rows:
            # Write header
            headers = list(rows[0].keys())
            for col_i, h in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col_i, value=h.replace("_", " ").title())
                cell.font = Font(bold=True)

            # Write rows
            for r_i, rec in enumerate(rows, start=2):
                for c_i, h in enumerate(headers, 1):
                    val = rec.get(h)
                    ws_data.cell(row=r_i, column=c_i, value=val)

        # Auto column width
        for column_cells in ws_data.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws_data.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 60)

        # =====================================================
        #  SHEET 2 : AI EXPLANATION
        # =====================================================
        ws_exp = wb.create_sheet("AI Explanation")
        ws_exp["A1"] = "User Query:"
        ws_exp["A1"].font = Font(bold=True)

        ws_exp["B1"] = log.query_text
        ws_exp["A3"] = "AI Explanation:"
        ws_exp["A3"].font = Font(bold=True)

        ws_exp["B3"] = explanation
        ws_exp["B3"].alignment = Alignment(wrap_text=True)

        ws_exp.column_dimensions["A"].width = 25
        ws_exp.column_dimensions["B"].width = 100

        # =====================================================
        #  SHEET 3 : CHART (Optional)
        # =====================================================
        try:
            chart_data = rows

            if rows and isinstance(rows, list):
                df = pd.DataFrame(rows)
                num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
                text_cols = [c for c in df.columns if df[c].dtype == "object"]

                if num_cols and text_cols:
                    x_col = text_cols[0]
                    df["x_axis"] = df[x_col].astype(str)

                    # Create plot
                    fig, ax = plt.subplots(figsize=(10, 4))

                    for y in num_cols:
                        ax.plot(df["x_axis"], df[y], marker="o", label=y)

                    plt.xticks(rotation=45, ha="right")
                    plt.tight_layout()

                    buf = io.BytesIO()
                    plt.savefig(buf, format="png", dpi=150)
                    buf.seek(0)

                    ws_chart = wb.create_sheet("Chart")
                    img = XLImage(buf)
                    img.anchor = "A1"
                    ws_chart.add_image(img)
                    plt.close()

        except Exception as e:
            _logger.warning("Chart creation failed for XLSX: %s", e)

        # =====================================================
        #  RETURN EXCEL
        # =====================================================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return request.make_response(
            output.getvalue(),
            headers=[
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="AI_School_Report_{search_log_id}.xlsx"')
            ]
        )
